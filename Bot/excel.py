import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from sqlighter import SQLighter

def write_messages():
	with SQLighter() as db:
		tmp = db.get_ru_messages()
		file_name = "../Texts.xlsx"
		
		wb = Workbook()
		ws1 = wb.active
		ws1.title = "Messages"
		ws1.append(["id", "step", "lang", "text_ru"])
		for button in tmp:
			ws1.append(button)
		wb.save(filename=file_name)


def save_users_to_file(filename):
	with SQLighter() as db:
		users = db.get_users_for_excel_file()
		wb = Workbook()
		ws1 = wb.active
		ws1.title = "Users"
		ws1.append(["user_id", "fullname", "username", "phone", "lang", "is_admin", "is_banned", "stopped_bot", "registration_date"])
		for user in users:
			ws1.append(user)
		wb.save(filename=filename)

	